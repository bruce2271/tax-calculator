"""Export a computed return to an Excel workbook.

The JSON save file reconstructs the return inside this app. It is not a working
paper — nobody reviews a return by reading JSON. A reviewer needs the figures
laid out by schedule, labelled with the IRS line they belong to, and with the
inputs sitting alongside the derived amounts so the arithmetic can be followed.

That is what this produces: one sheet per schedule, three layers on every sheet
(input, derived, result), and an IRS form and line reference wherever the field
maps to one.

Uses openpyxl directly rather than pandas: the dependency is already pinned for
the trial balance importer, and writing cells directly keeps control of the
number formats, which matter more here than dataframe convenience.
"""

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from calculators.tax_lines import TAX_LINES

# field name -> (form, line, label). Built once; TAX_LINES is a module constant.
_BY_FIELD = {row[4]: (row[1], row[2], row[3]) for row in TAX_LINES}

ACCOUNTING_FORMAT = '#,##0.00;(#,##0.00);"—"'
PERCENT_FORMAT = '0.00%'

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
_SECTION_FONT = Font(bold=True, size=10)
_SECTION_FILL = PatternFill("solid", fgColor="D9E2F3")
_TOTAL_FONT = Font(bold=True)
_TOP_BORDER = Border(top=Side(style="thin"))

_COLUMNS = ("Layer", "Form", "Line", "Item", "Amount")

# Keys whose value is a rate rather than a currency amount.
_RATE_KEYS = {"effective_rate", "rate", "effective_limit"}


def _is_number(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def flatten(node, prefix=""):
    """Walk a nested result dict into (path, value) pairs, numbers only.

    Nested dicts are joined with " > " so a reviewer can see which block a figure
    came from — "credits > foreign_tax_credit" rather than a bare label that
    appears under three different schedules.
    """
    rows = []
    for key, value in node.items():
        path = f"{prefix} > {key}" if prefix else str(key)
        if isinstance(value, dict):
            rows.extend(flatten(value, path))
        elif _is_number(value):
            rows.append((path, value))
    return rows


def humanise(path):
    """'tax > credits > foreign_tax_credit' -> 'Foreign tax credit'.

    Only the leaf is turned into a label; the parent path is already carried by
    the sheet the row sits on.
    """
    leaf = path.split(" > ")[-1]
    return leaf.replace("_", " ").strip().capitalize()


def line_reference(path):
    """(form, line) for a field that maps to an IRS line, else ("", "")."""
    leaf = path.split(" > ")[-1]
    form, line, _label = _BY_FIELD.get(leaf, ("", "", ""))
    return form, line


def _write_header(ws, title, subtitle=None):
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13)
    row = 2
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(italic=True, size=9, color="595959")
        row = 3
    header_row = row + 1
    for col, name in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    return header_row + 1


def _write_rows(ws, start_row, rows, layer):
    """rows: iterable of (path, value). Returns the next free row."""
    r = start_row
    for path, value in rows:
        form, line = line_reference(path)
        ws.cell(row=r, column=1, value=layer)
        ws.cell(row=r, column=2, value=form)
        ws.cell(row=r, column=3, value=line)
        ws.cell(row=r, column=4, value=humanise(path))
        cell = ws.cell(row=r, column=5, value=value)
        leaf = path.split(" > ")[-1]
        cell.number_format = PERCENT_FORMAT if leaf in _RATE_KEYS else ACCOUNTING_FORMAT
        if leaf.startswith("total_") or leaf.startswith("taxable_income"):
            ws.cell(row=r, column=4).font = _TOTAL_FONT
            cell.font = _TOTAL_FONT
            cell.border = _TOP_BORDER
        r += 1
    return r


def _autosize(ws):
    widths = {1: 10, 2: 10, 3: 8, 4: 46, 5: 16}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A5"


def _disclaimer_sheet(wb, tax_year):
    ws = wb.create_sheet("Read me", 0)
    ws["A1"] = "Educational tool — not tax advice"
    ws["A1"].font = Font(bold=True, size=14, color="C00000")
    lines = [
        "",
        f"Tax year: {tax_year}",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "",
        "This workbook was produced by an educational model of the U.S. corporate income",
        "tax return. It was not prepared or reviewed by a CPA or an enrolled agent, it does",
        "not produce a filable return, and it simplifies or omits many rules that apply to a",
        "real taxpayer.",
        "",
        "Do not rely on it to compute an actual tax liability. Verify every figure against the",
        "current IRS forms and instructions, and consult a qualified tax professional.",
        "",
        "How to read the sheets",
        "  Layer   Input    — a figure that was entered",
        "          Derived  — an intermediate amount the model computed",
        "          Result   — a figure that lands on the return",
        "  Form / Line      — the IRS form and line the item reports to, where one applies",
        "",
        "Inputs are included deliberately: a working paper that shows only answers cannot be",
        "reviewed. The arithmetic has to be followable from what was entered.",
    ]
    for i, text in enumerate(lines, start=2):
        ws.cell(row=i, column=1, value=text)
        if text.strip() == "How to read the sheets":
            ws.cell(row=i, column=1).font = _SECTION_FONT
    ws.column_dimensions["A"].width = 92
    return ws


def build_workbook(inputs, results, tax_year=None, extra_sheets=None):
    """Return the workbook as bytes, ready for st.download_button.

    inputs        flat dict of entered figures
    results       nested dict returned by calculate_1120
    extra_sheets  optional {sheet name: nested dict} for additional schedules
    """
    wb = Workbook()
    wb.remove(wb.active)

    _disclaimer_sheet(wb, tax_year)

    # ── Summary: the figures a reviewer looks for first ──────────────────────
    ws = wb.create_sheet("Summary")
    row = _write_header(ws, "Summary", "The figures that carry to the return")
    summary_paths = [
        ("income", "total_income"),
        ("deductions", "total_deductions"),
        ("taxable_income", "taxable_income"),
        ("tax", "regular_tax"),
        ("tax", "tax_after_credits"),
        ("tax", "total_federal_tax"),
    ]
    picked = []
    for block, key in summary_paths:
        value = results.get(block, {}).get(key)
        if _is_number(value):
            picked.append((f"{block} > {key}", value))
    row = _write_rows(ws, row, picked, "Result")
    _autosize(ws)

    # ── One sheet per top-level block of the result ──────────────────────────
    for block, content in results.items():
        if not isinstance(content, dict):
            continue
        ws = wb.create_sheet(str(block).replace("_", " ").title()[:31])
        row = _write_header(ws, str(block).replace("_", " ").title())
        rows = flatten(content)
        layer = "Result" if block in ("taxable_income", "tax") else "Derived"
        _write_rows(ws, row, rows, layer)
        _autosize(ws)

    # ── Inputs last: the audit trail, not the headline ───────────────────────
    ws = wb.create_sheet("Inputs")
    row = _write_header(ws, "Inputs", "Everything entered, so the arithmetic can be followed")
    entered = sorted((k, v) for k, v in inputs.items() if _is_number(v))
    _write_rows(ws, row, entered, "Input")
    _autosize(ws)

    for name, content in (extra_sheets or {}).items():
        ws = wb.create_sheet(str(name)[:31])
        row = _write_header(ws, str(name))
        _write_rows(ws, row, flatten(content), "Derived")
        _autosize(ws)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
