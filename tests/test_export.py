from io import BytesIO

import pytest
from openpyxl import load_workbook

from calculators.export import (build_workbook, flatten, humanise,
                                line_reference)


# ── flatten ───────────────────────────────────────────────────────────────────

def test_flatten_walks_nested_dicts():
    node = {"tax": {"credits": {"foreign_tax_credit": 1200.0}}}
    assert flatten(node) == [("tax > credits > foreign_tax_credit", 1200.0)]


def test_flatten_keeps_zero_but_drops_non_numbers():
    # Zero is a real answer on a tax return — "no deduction" is not the same as
    # "not computed", so it has to survive.
    node = {"a": 0, "b": "SL 39yr", "c": None, "d": 5.5}
    assert flatten(node) == [("a", 0), ("d", 5.5)]


def test_flatten_drops_booleans():
    # bool is a subclass of int; an election flag is not an amount.
    assert flatten({"elected": True, "amount": 10}) == [("amount", 10)]


def test_flatten_handles_empty():
    assert flatten({}) == []


# ── labelling ─────────────────────────────────────────────────────────────────

def test_humanise_uses_the_leaf_only():
    assert humanise("tax > credits > foreign_tax_credit") == "Foreign tax credit"


def test_line_reference_finds_a_mapped_field():
    form, line = line_reference("income > gross_receipts_book")
    assert (form, line) == ("1120", "1a")


def test_line_reference_is_blank_for_unmapped_fields():
    assert line_reference("tax > some_internal_intermediate") == ("", "")


# ── workbook ──────────────────────────────────────────────────────────────────

@pytest.fixture
def workbook():
    inputs = {"gross_receipts_book": 1_000_000, "salaries_book": 250_000,
              "entity_name": "ACME Corp"}
    results = {
        "income": {"gross_revenue": 1_000_000, "total_income": 1_050_000},
        "deductions": {"cogs": 400_000, "total_deductions": 650_000},
        "taxable_income": {"before_nol": 400_000, "taxable_income": 400_000},
        "tax": {"regular_tax": 84_000, "effective_rate": 0.21,
                "credits": {"foreign_tax_credit": 1_200, "total_credits": 1_200},
                "tax_after_credits": 82_800, "total_federal_tax": 82_800},
    }
    raw = build_workbook(inputs, results, tax_year=2024)
    return load_workbook(BytesIO(raw))


def test_read_me_sheet_comes_first(workbook):
    # A disclaimer nobody sees is not a disclaimer.
    assert workbook.sheetnames[0] == "Read me"


def test_every_result_block_gets_a_sheet(workbook):
    for name in ("Summary", "Income", "Deductions", "Taxable Income", "Tax", "Inputs"):
        assert name in workbook.sheetnames


def test_summary_carries_the_headline_figures(workbook):
    values = [c.value for row in workbook["Summary"].iter_rows() for c in row]
    assert 400_000 in values      # taxable income
    assert 82_800 in values       # total federal tax


def test_inputs_sheet_holds_entered_figures_only(workbook):
    values = [c.value for row in workbook["Inputs"].iter_rows() for c in row]
    assert 1_000_000 in values
    assert "ACME Corp" not in values      # text input, not an amount


def test_inputs_are_labelled_with_their_irs_line(workbook):
    rows = list(workbook["Inputs"].iter_rows(values_only=True))
    gross = [r for r in rows if r[3] == "Gross receipts book"]
    assert gross and gross[0][1] == "1120" and gross[0][2] == "1a"


def test_nested_credits_reach_the_tax_sheet(workbook):
    labels = [r[3] for r in workbook["Tax"].iter_rows(values_only=True)]
    assert "Foreign tax credit" in labels


def test_rates_are_formatted_as_percentages(workbook):
    for row in workbook["Tax"].iter_rows():
        if row[3].value == "Effective rate":
            assert "%" in row[4].number_format
            break
    else:
        pytest.fail("effective rate did not reach the Tax sheet")


def _layers(ws):
    """Column A below the header row. The title and subtitle also sit in A."""
    rows = list(ws.iter_rows(values_only=True))
    header = next(i for i, r in enumerate(rows) if r[0] == "Layer")
    return {r[0] for r in rows[header + 1:] if r[0]}


def test_layer_column_separates_input_from_result(workbook):
    assert _layers(workbook["Inputs"]) == {"Input"}
    assert _layers(workbook["Tax"]) == {"Result"}
    assert _layers(workbook["Deductions"]) == {"Derived"}


def test_extra_sheets_are_appended():
    raw = build_workbook({}, {}, tax_year=2024,
                         extra_sheets={"Schedule L": {"total_assets": 5_000_000}})
    wb = load_workbook(BytesIO(raw))
    assert "Schedule L" in wb.sheetnames
    values = [c.value for row in wb["Schedule L"].iter_rows() for c in row]
    assert 5_000_000 in values


def test_empty_return_still_produces_a_readable_file():
    # A blank return should not crash the export; the reviewer still gets the
    # disclaimer and an empty shell rather than an error.
    wb = load_workbook(BytesIO(build_workbook({}, {}, tax_year=2024)))
    assert wb.sheetnames[0] == "Read me"
    assert "Inputs" in wb.sheetnames
